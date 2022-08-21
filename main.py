from openpyxl import load_workbook

# Загружаем файл и переходим на первую страницу
wb = load_workbook('1.xlsx')
sheet = wb.get_sheet_by_name(wb.get_sheet_names()[0])
f = list(sheet.values)
# Удалим все нечетные числа
f = [[None if f[i][j] % 2 else f[i][j] for i in range(sheet.max_column)] for j in range(sheet.max_row)]
print(*f, sep='\n')

for i in range(1, len(f) + 1):
    for j in range(1, len(f[0]) + 1):
        sheet.cell(row=i, column=j).value = f[i-1][j-1]
wb.save('2.xlsx')
